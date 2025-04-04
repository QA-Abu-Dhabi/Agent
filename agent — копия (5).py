import pdfplumber
import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Файл для сохранения результатов
excel_filename = "result.xlsx"

# Функция для поиска PDF-файлов по шаблону
def find_pdf_files(pattern):
    file_pattern = os.path.join(os.getcwd(), '**', pattern)  # Шаблон поиска
    all_matches = glob.glob(file_pattern, recursive=True)  # Поиск всех совпадений
    return [f for f in all_matches if os.path.isfile(f)]  # Оставляем только файлы

# Находим файлы
pdf_files_kc2 = find_pdf_files('Печатная форма*КС2*.pdf')
pdf_files_oa = find_pdf_files('Печатная форма ОА*.pdf')
pdf_files_sf = find_pdf_files('Печатная форма Счет-фактура*.pdf')

# Список для хранения результатов
results = []

# Функция для извлечения суммы из КС2
def extract_kc2_sum(pdf_filename):
    with pdfplumber.open(pdf_filename) as pdf:
        for page in reversed(pdf.pages):  # Перебираем страницы с конца
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if row and row[0] and ("всего по акту" in row[0].lower()):
                        if len(row) > 7 and row[7]:  # Проверяем 8-й столбец
                            raw_value = row[7].replace("\u00A0", "").replace(" ", "")  # Убираем пробелы
                            try:
                                return round(float(raw_value.replace(",", ".")), 2)
                            except ValueError:
                                pass  # Если не число, продолжаем искать в 9 столбце
                            
                        if len(row) > 8 and row[8]:  # Проверяем 9-й столбец
                            raw_value = row[8].replace("\u00A0", "").replace(" ", "")  # Убираем пробелы
                            try:
                                return round(float(raw_value.replace(",", ".")), 2)
                            except ValueError:
                                return raw_value  # Если не число — оставляем как есть
    return "Не найдено"

# Функция для извлечения значений из 4-й строки 7 и 11 столбцов в OA-файлах
def extract_oa_values(pdf_filename):
    with pdfplumber.open(pdf_filename) as pdf:
        for page in pdf.pages:  # OA-файлы читаем с начала
            tables = page.extract_tables()
            for table in tables:
                if len(table) > 3:  # Проверяем, есть ли хотя бы 4 строки
                    row = table[3]  # Берём 4-ю строку (индекс 3)
                    if len(row) > 6 and len(row) > 10:  # Проверяем 7-й и 11-й столбцы
                        value_7 = row[6].replace("\u00A0", "").replace(" ", "").replace(",", ".") if row[6] else "Не найдено"
                        value_11 = row[10].replace("\u00A0", "").replace(" ", "").replace(",", ".") if row[10] else "Не найдено"
                        try:
                            value_7 = round(float(value_7), 2)
                        except ValueError:
                            pass  # Оставляем как есть
                        try:
                            value_11 = round(float(value_11), 2)
                        except ValueError:
                            pass  # Оставляем как есть
                        return value_7, value_11  # Возвращаем оба значения
    return "Не найдено", "Не найдено"

# Функция для обработки Счет-фактуры
def extract_sf_values(pdf_filename):
    with pdfplumber.open(pdf_filename) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if len(table) > 3:  # Должно быть минимум 4 строки
                    row = table[3]  # 4-я строка (индекс 3)
                    if len(row) > 7:  # Должно быть минимум 8 столбцов
                        value_8 = row[7].replace("\u00A0", "").replace(" ", "").replace(",", ".") if row[7] else "Не найдено"
                        try:
                            value_8 = round(float(value_8), 2)
                        except ValueError:
                            pass  # Оставляем как есть

                        # Определяем тип
                        sf_type = "СФ АВ" if len(row) > 1 and row[1] and "Агентское вознаграждение" in row[1] else "СФ СМР"
                        
                        return sf_type, value_8
    return "Не найдено", "Не найдено"

# Обрабатываем КС2-файлы
for pdf_filename in pdf_files_kc2:
    extracted_sum = extract_kc2_sum(pdf_filename)
    results.append({
        "Тип": "КС2",
        "Файл": os.path.basename(pdf_filename),
        "Сумма": extracted_sum
    })

# Обрабатываем OA-файлы
for pdf_filename in pdf_files_oa:
    value_7, value_11 = extract_oa_values(pdf_filename)

    # Каждое значение в отдельной строке
    results.append({
        "Тип": "OA СМР",
        "Файл": os.path.basename(pdf_filename),
        "Сумма": value_7
    })
    results.append({
        "Тип": "OA АВ",
        "Файл": os.path.basename(pdf_filename),
        "Сумма": value_11
    })

# Обрабатываем Счет-фактуру
for pdf_filename in pdf_files_sf:
    sf_type, value_8 = extract_sf_values(pdf_filename)
    results.append({
        "Тип": sf_type,
        "Файл": os.path.basename(pdf_filename),
        "Сумма": value_8
    })

# Сохраняем в Excel
df = pd.DataFrame(results)
df.to_excel(excel_filename, index=False, engine="openpyxl")

# открываем сохраненный файл для добавления формулы
wb = load_workbook(excel_filename)
ws = wb.active

# Настройка ширины столбцов
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 9

# Выравнивание по правому краю
for row in ws['D']:
    row.alignment = Alignment(horizontal='right')

# Добавляем формулу
ws['D2'] = 'OA СМР = СФ СМР'
ws['D3'] = 'OA АВ = СФ АВ'
ws['D4'] = 'OA СМР = сумме КС2'
ws['E2'] = '=VLOOKUP("OA СМР",A1:C20,3,0)=VLOOKUP("СФ СМР",A1:C20,3,0)' #=ВПР("OA СМР";A1:C20;3;0)=ВПР("СФ СМР";A1:C20;3;0)
ws['E3'] = '=VLOOKUP("OA АВ",A1:C20,3,0)=VLOOKUP("СФ АВ",A1:C20,3,0)'
ws['E4'] = '=VLOOKUP("OA СМР",A1:C20,3,0)=SUMIF(A1:A20,"КС2",C1:C20)'

# Сохраняем изменения
wb.save(excel_filename)

print(f"Обработано файлов: {len(results)}. Результаты сохранены в {excel_filename}.")