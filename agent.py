# Для корректной компиляции кода в exe файл необходимо запустить следующую команду:
# pyinstaller --onefile --windowed --icon=favicon.ico --add-data "favicon.ico;." agent.py

import sys
import os
import glob
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QMessageBox, QLabel
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt


class PDFProcessor(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()
        self.folder_path = ""

    def init_ui(self):
        layout = QVBoxLayout()
        
        # Добавляем описание программы в QLabel
        self.label_description = QLabel(
            "Данная программа предназначена для извлечения данных\n"
            "из Отчета агента, Формы КС-2, Счета-фактуры\n\n"
            "Для распознавания необходимо выбрать папку с файлами в PDF формате."
        )
        self.label_description.setWordWrap(True)  # Разрешаем перенос строк
        self.label_description.setAlignment(Qt.AlignLeft)  # Выравниваем текст по левому краю
        layout.addWidget(self.label_description)  # Добавляем в макет

        # Кнопка для выбора папки
        self.button_select_folder = QPushButton('Выбрать папку', self)
        self.button_select_folder.clicked.connect(self.select_folder)
        self.button_select_folder.setFixedSize(300, 70)
        layout.addWidget(self.button_select_folder, alignment=Qt.AlignCenter)

        # Метка для отображения пути к выбранной папке
        self.label_folder_path = QLabel("Папка не выбрана", self)
        self.label_folder_path.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label_folder_path)
        
        # Кнопка для запуска обработки
        self.button_process_files = QPushButton('Запустить обработку', self)
        self.button_process_files.clicked.connect(self.process_files)
        self.button_process_files.setFixedSize(300, 70)
        self.button_process_files.setEnabled(False)  # Отключена, пока не выбрана папка
        layout.addWidget(self.button_process_files, alignment=Qt.AlignCenter)
        
        layout.addStretch() # Прижимаем текст к нижнему краю
        self.label_author = QLabel("© 2025 Автор: Абубакиров Ильмир Иргалиевич", self)
        self.label_author.setAlignment(Qt.AlignRight)  # Выравнивание справа
        layout.addWidget(self.label_author)
        self.label_author.setStyleSheet("font-size: 12px; color: gray;")
        
        if getattr(sys, 'frozen', False):  # Проверяем, запущен ли exe-файл
            base_path = sys._MEIPASS  # Временная папка PyInstaller
        else:
            base_path = os.path.dirname(__file__)  # Обычный путь для скрипта
            
        icon_path = os.path.join(base_path, "favicon.ico")
        self.setWindowIcon(QIcon(icon_path))
        
        self.setLayout(layout)
        self.setWindowTitle('Обработка PDF файлов')
        self.setGeometry(300, 200, 600, 500)
        
        layout.setSpacing(40)

    def select_folder(self):
        # Открываем диалог для выбора папки
        folder_path = QFileDialog.getExistingDirectory(self, 'Выберите папку')
        if folder_path:
            self.folder_path = folder_path
            self.label_folder_path.setText(f"Выбрана папка:\n{folder_path}")
            self.button_process_files.setEnabled(True)  # Включаем кнопку распознавания
            #QMessageBox.information(self, 'Выбрана папка', f'Выбрана папка: {self.folder_path}')
        else:
            QMessageBox.warning(self, 'Ошибка', 'Папка не выбрана!')

    def process_files(self):
        if not self.folder_path:
            QMessageBox.warning(self, 'Ошибка', 'Не выбрана папка!')
            return

        # Функции обработки PDF файлов
        pdf_files_kc2 = self.find_pdf_files('Печатная форма*КС2*.pdf', '*КС-2*.pdf')
        pdf_files_oa = self.find_pdf_files('*ОА*.pdf')
        pdf_files_sf = self.find_pdf_files('Печатная форма Счет-фактура*.pdf', '*№ГИ*.pdf')

        results = []

        # Обрабатываем файлы
        results += self.process_kc2_files(pdf_files_kc2)
        results += self.process_oa_files(pdf_files_oa)
        results += self.process_sf_files(pdf_files_sf)

        # Сохраняем в Excel
        self.save_to_excel(results)

    def find_pdf_files(self, pattern1, pattern2=None):
        file_pattern1 = os.path.join(self.folder_path, '**', pattern1)
        all_matches = glob.glob(file_pattern1, recursive=True)

        if pattern2:  # Если указан второй шаблон, ищем и по нему
            file_pattern2 = os.path.join(self.folder_path, '**', pattern2)
            all_matches += glob.glob(file_pattern2, recursive=True)

        return [f for f in all_matches if os.path.isfile(f)]

    def process_kc2_files(self, pdf_files):
        # Логика обработки КС2 файлов
        results = []
        for pdf_filename in pdf_files:
            extracted_sum = self.extract_kc2_sum(pdf_filename)
            results.append({
                "Тип": "КС2",
                "Файл": os.path.basename(pdf_filename),
                "Сумма": extracted_sum
            })
        return results

    def process_oa_files(self, pdf_files):
        # Логика обработки OA файлов
        results = []
        for pdf_filename in pdf_files:
            value_7, value_11 = self.extract_oa_values(pdf_filename)
            results.append({
                "Тип": "OA СМР",
                "Файл": os.path.basename(pdf_filename),
                "Сумма": value_7
            })
            results.append({
                "Тип": "OA АВ",
                "Файл": os.path.basename(pdf_filename),
                "Сумма": value_11
            })
        return results

    def process_sf_files(self, pdf_files):
        # Логика обработки Счет-фактур
        results = []
        for pdf_filename in pdf_files:
            sf_type, value_8 = self.extract_sf_values(pdf_filename)
            results.append({
                "Тип": sf_type,
                "Файл": os.path.basename(pdf_filename),
                "Сумма": value_8
            })
        return results

    def extract_kc2_sum(self, pdf_filename):
        print(f"\nОткрытие файла: {pdf_filename}")
        with pdfplumber.open(pdf_filename) as pdf:
            total_pages = len(pdf.pages)
            print(f"Всего страниц в PDF: {total_pages}")
            
            # Сначала ищем строку со словами "всего по акту"
            for i, page in enumerate(reversed(pdf.pages), start=1):
                page_number = total_pages - i + 1
                print(f"\nОбработка страницы {page_number} (с конца)")

                tables = page.extract_tables()
                print(f"  Найдено таблиц: {len(tables)}")

                for t_index, table in enumerate(tables):
                    print(f"    Обработка таблицы {t_index + 1}")
                    for r_index, row in enumerate(table):
                        if not row:
                            continue
                        if row[0] and "всего по акту" in row[0].lower():
                            print(f"      Найдена строка со словами 'всего по акту': {row}")
                            for col_index in [7, 8]:
                                if len(row) > col_index and row[col_index]:
                                    raw_value = row[col_index].replace("\u00A0", "").replace(" ", "")
                                    print(f"        Попытка извлечь значение из столбца {col_index + 1}: {raw_value}")
                                    try:
                                        number = round(float(raw_value.replace(",", ".")), 2)
                                        print(f"        Успешно извлечено число: {number}")
                                        return number
                                    except ValueError:
                                        print("        Ошибка преобразования в число")
                                        continue
            
            print("\nНе удалось найти строку 'всего по акту'. Переход к альтернативному способу...")

            # Альтернатива — последняя строка последней страницы
            def try_extract_from_page(page, page_index):
                tables = page.extract_tables()
                print(f"\nСтраница {page_index + 1}: найдено таблиц: {len(tables)}")

                if not tables:
                    return None  # таблиц нет — вернём None

                last_table = tables[-1]
                last_row = last_table[-1]
                print(f"Последняя строка таблицы: {last_row}")

                for col_index, col in enumerate(reversed(last_row), start=1):
                    if col:
                        raw_value = col.replace("\u00A0", "").replace(" ", "")
                        print(f"  Проверка значения (с конца, столбец -{col_index}): {raw_value}")
                        try:
                            number = round(float(raw_value.replace(",", ".")), 2)
                            print(f"  Успешно извлечено число: {number}")
                            return number
                        except ValueError:
                            print("  Не удалось преобразовать в число.")
                            continue
                return None

            # Пробуем на последней странице
            last_page_index = total_pages - 1
            result = try_extract_from_page(pdf.pages[last_page_index], last_page_index)

            # Если не получилось — пробуем на предпоследней
            if result is None and total_pages >= 2:
                print("\nНа последней странице таблиц нет или значение не найдено. Пробуем на предпоследней...")
                prev_page_index = total_pages - 2
                result = try_extract_from_page(pdf.pages[prev_page_index], prev_page_index)

            if result is not None:
                return result

        print("Итог: число не найдено.")
        return "Не найдено"

    def extract_oa_values(self, pdf_filename):
        with pdfplumber.open(pdf_filename) as pdf:
            for page_num, page in enumerate(reversed(pdf.pages), start=1):
                tables = page.extract_tables()
                if not tables:
                    continue  # Пропускаем страницу, если таблиц нет
                
                for table_num, table in enumerate(tables, start=1):
                    for row_num, row in enumerate(table, start=1):
                        if any(cell and "всего:" in cell.lower() for cell in row if cell):  # Ищем слово "Всего:"
                            print(f"Найдена строка {row_num} на странице {len(pdf.pages) - page_num + 1}: {row}")

                            value_7 = row[6] if len(row) > 6 and row[6] else "Не найдено"
                            value_11 = row[10] if len(row) > 10 and row[10] else "Не найдено"

                            value_7 = value_7.replace("\u00A0", "").replace(" ", "").replace(",", ".")
                            value_11 = value_11.replace("\u00A0", "").replace(" ", "").replace(",", ".")

                            try:
                                value_7 = round(float(value_7), 2)
                            except ValueError:
                                value_7 = "Ошибка"
                            
                            try:
                                value_11 = round(float(value_11), 2)
                            except ValueError:
                                value_11 = "Ошибка"

                            return value_7, value_11  # Как только нашли строку, выходим

        return "Не найдено", "Не найдено"
    
    
    def extract_sf_values(self, pdf_filename):
        print(f"\nОткрытие файла: {pdf_filename}")
        with pdfplumber.open(pdf_filename) as pdf:
            total_pages = len(pdf.pages)
            print(f"Всего страниц: {total_pages}")

            for offset in [1, 2]:  # Последняя и предпоследняя страницы
                page_index = total_pages - offset
                if page_index < 0:
                    continue

                page = pdf.pages[page_index]
                print(f"\nПроверка страницы {page_index + 1}")

                tables = page.extract_tables()
                print(f"  Найдено таблиц: {len(tables)}")

                if not tables:
                    print("  Таблиц не найдено на странице.")
                    continue

                last_table = tables[-1]
                if not last_table or len(last_table) < 2:
                    print("  Таблица пуста или недостаточно строк.")
                    continue

                # Последняя строка — значение
                last_row = last_table[-1]
                # Предпоследняя строка — определение типа
                second_last_row = last_table[-2]

                print(f"  Последняя строка: {last_row}")
                print(f"  Предпоследняя строка: {second_last_row}")

                # Извлечение значения из столбца 8
                if len(last_row) > 7 and last_row[7]:
                    raw_value = last_row[7].replace("\u00A0", "").replace(" ", "").replace(",", ".")
                    try:
                        value_8 = round(float(raw_value), 2)
                    except ValueError:
                        value_8 = "Ошибка"
                    print(f"  Извлечённое значение из столбца 8: {value_8}")
                else:
                    value_8 = "Не найдено"
                    print("  Столбец 8 отсутствует или пуст.")

                # Определение типа по содержимому 2-го столбца предпоследней строки
                if len(second_last_row) > 1 and second_last_row[1] and "Агентское вознаграждение" in second_last_row[1]:
                    sf_type = "СФ АВ"
                else:
                    sf_type = "СФ СМР"
                print(f"  Тип СФ: {sf_type}")

                return sf_type, value_8

        print("  Не удалось извлечь данные.")
        return "Не найдено", "Не найдено"

    def save_to_excel(self, results):
        # Сохраняем результаты в Excel
        df = pd.DataFrame(results)
        excel_filename = "result.xlsx"
        df.to_excel(excel_filename, index=False, engine="openpyxl")

        # Открываем и добавляем формулы в файл
        wb = load_workbook(excel_filename)
        ws = wb.active
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 9

        for row in ws['D']:
            row.alignment = Alignment(horizontal='right')

        ws['D2'] = 'OA СМР = СФ СМР'
        ws['D3'] = 'OA АВ = СФ АВ'
        ws['D4'] = 'OA СМР = сумме КС2'
        ws['E2'] = '=VLOOKUP("OA СМР",A1:C500,3,0)=VLOOKUP("СФ СМР",A1:C500,3,0)'
        ws['E3'] = '=VLOOKUP("OA АВ",A1:C500,3,0)=VLOOKUP("СФ АВ",A1:C500,3,0)'
        ws['E4'] = '=VLOOKUP("OA СМР",A1:C500,3,0)=SUMIF(A1:A500,"КС2",C1:C500)'
        
        # Создаем стиль для чисел с разделителем
        num_style = NamedStyle(name='num_style', number_format='#,##0.00')
        
        # Применяем стиль к ячейкам столбца С
        for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, max_row=500):
            for cell in row:
                cell.style = num_style

        wb.save(excel_filename)

        QMessageBox.information(self, 'Готово', f'Обработано файлов: {len(results)}. Результаты сохранены в {excel_filename}.')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = PDFProcessor()
    window.show()
    sys.exit(app.exec_())